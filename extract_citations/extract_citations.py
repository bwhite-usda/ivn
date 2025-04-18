# Filename: extract_citations.py


import requests
import PyPDF2
import re
import os
import tempfile
import time
from urllib3.util.retry import Retry
from requests.adapters import HTTPAdapter
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter


def sanitize_text(text):
    return re.sub(r"[\r\n]+", " ", text).strip()


def clean_citation(citation):
    # Normalize U.S. Code references
    citation = re.sub(
        r"\b(\d+)\s*(U\.S\.C\.|USC|U\.S\. Code)\s*§?\s*(\d+(\.\d+)*([a-zA-Z0-9()]*)?)",
        r"\1 USC \3",
        citation,
        flags=re.IGNORECASE
    )


    # Normalize CFR references
    citation = re.sub(
        r"\b(\d+)\s*(C\.F\.R\.|CFR|Code of Federal Regulations)\s*§?\s*(\d+(\.\d+)*([a-zA-Z0-9()]*)?)",
        r"\1 CFR \3",
        citation,
        flags=re.IGNORECASE
    )


    # Normalize Executive Order references
    citation = re.sub(
        r"\b(E\.O\.|Executive\s*Order)\s*(\d+)",
        r"Executive Order \2",
        citation,
        flags=re.IGNORECASE
    )


    citation = re.sub(
        r"\bEO\s+(\d+)\b",
        r"Executive Order \1",
        citation,
        flags=re.IGNORECASE
    )


    # Normalize Public Laws
    citation = re.sub(
        r"\b(Public\s+Law|P\.L\.)\s*(\d{1,3}[-–]\d{1,4})",
        r"Public Law \2",
        citation,
        flags=re.IGNORECASE
    )


    # Normalize Acts
    citation = re.sub(
        r"\bAct\s+of\s+(\d{4})",
        r"Act of \1",
        citation,
        flags=re.IGNORECASE
    )


    # Normalize Title references
    citation = re.sub(
        r"\bTitle\s+(\d+)",
        r"Title \1",
        citation,
        flags=re.IGNORECASE
    )


    return citation


def get_browser_headers():
    return {
        "User-Agent": (
            "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
            "AppleWebKit/537.36 (KHTML, like Gecko) "
            "Chrome/122.0.0.0 Safari/537.36"
        ),
        "Accept": "application/pdf",
        "Connection": "keep-alive"
    }


def download_pdf(url):
    try:
        session = requests.Session()
        retries = Retry(
            total=5,
            backoff_factor=5,
            status_forcelist=[500, 502, 503, 504],
            raise_on_status=False,
        )
        session.mount('https://', HTTPAdapter(max_retries=retries))


        response = session.get(url, headers=get_browser_headers(), stream=True, timeout=60)
        response.raise_for_status()


        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".pdf")
        for chunk in response.iter_content(chunk_size=1024):
            temp_file.write(chunk)
        temp_file.close()


        print(f"Downloaded {url}")
        return temp_file.name
    except Exception as e:
        print(f"Failed to download {url}: {e}")
        with open("failed_downloads.txt", "a") as f:
            f.write(url + "\n")
        return None


def extract_toc(reader):
    toc = []
    toc_pattern = r"(?P<heading>.+?)\s+(\d+)"
    for page_num, page in enumerate(reader.pages[:10]):
        text = page.extract_text()
        if text and "Table of Contents" in text:
            matches = re.findall(toc_pattern, text)
            for match in matches:
                heading = sanitize_text(match[0])
                page_start = int(match[1])
                toc.append((heading, page_start))
    return toc


def infer_section_name(toc, page_num, context, page_text):
    if toc:
        for i, (section, start_page) in enumerate(toc):
            if i + 1 < len(toc) and toc[i + 1][1] > page_num >= start_page:
                return section
            elif i == len(toc) - 1 and page_num >= start_page:
                return section
    lines = page_text.splitlines()
    context_start = page_text.find(context)
    for i in range(len(lines) - 1, -1, -1):
        if len(lines[i].strip()) > 0 and lines[i].strip() in page_text[:context_start]:
            return sanitize_text(lines[i])
    return "Unknown Section"


def extract_us_code_citations(pdf_path, url):
    try:
        with open(pdf_path, 'rb') as file:
            reader = PyPDF2.PdfReader(file)
            toc = extract_toc(reader)
            num_pages = len(reader.pages)
            citations = []


            citation_pattern = (
                r"\b(\d+)\s*(U\.S\.C\.|USC|U\.S\. Code)\s*§?\s*(\d+(\.\d+)*([a-zA-Z0-9]*)?)|"
                r"\b(\d+)\s*(C\.F\.R\.|CFR|Code of Federal Regulations)\s*§?\s*(\d+(\.\d+)*([a-zA-Z0-9]*)?)|"
                r"(E\.O\.|Executive\s*Order)\s*(\d+)|"
                r"\bEO\s+(\d+)\b|"
                r"\bPublic\s+Law\s+\d{1,3}[-–]\d{1,4}\b|"
                r"\bP\.L\.\s*\d{1,3}[-–]\d{1,4}\b|"
                r"\bAct\s+of\s+\d{4}\b|"
                r"\bTitle\s+\d+\b"
            )


            for page_num in range(num_pages):
                page = reader.pages[page_num]
                text = page.extract_text()
                if text:
                    matches = re.finditer(citation_pattern, text, re.IGNORECASE)
                    for match in matches:
                        citation_text = match.group(0)
                        citation = clean_citation(citation_text)
                        start, end = match.start(), match.end()
                        context = sanitize_text(text[max(0, start - 100):min(len(text), end + 100)])
                        section_name = infer_section_name(toc, page_num + 1, context, text)
                        citation_page_url = f"{url}#page={page_num + 1}"
                        citations.append((citation, citation_page_url, section_name, context, url))
        return citations
    except Exception as e:
        print(f"Error processing {pdf_path}: {e}")
        return []


def process_url(url):
    temp_file = download_pdf(url)
    if not temp_file:
        return []
    try:
        return extract_us_code_citations(temp_file, url)
    finally:
        os.remove(temp_file)


def save_to_excel(data, filename="extracted_citations.xlsx"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Citation", "Citation Page", "Inferred Section Name", "Context", "URL"])


    for row in data:
        sanitized_row = [sanitize_text(str(cell)) for cell in row]
        sanitized_row[0] = clean_citation(sanitized_row[0])
        sheet.append(sanitized_row)


    for col in range(1, sheet.max_column + 1):
        sheet.column_dimensions[get_column_letter(col)].width = 20


    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
        for cell in row:
            cell.hyperlink = cell.value
            cell.style = "Hyperlink"


    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=5, max_col=5):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True)


    workbook.save(filename)
    print(f"Saved data to {filename}")


def main():
    url_list = [        
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/10240.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-08/10000.1_0.pdf",
        "https://www.fsis.usda.gov/policy/fsis-directives/10010.1.pdf",
        "https://www.fsis.usda.gov/policy/fsis-directives/10010.2",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-08/10100.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2022-05/4338.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-08/10010.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1010.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/10010.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1020.4.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/10210.5.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/10230.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/10230.6.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/10240.4.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/10240.5.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/10250.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/10250.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/10300.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/10310.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1040.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/10400.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1045.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1050.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1060.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1070.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/10800.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/10800.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/10800.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1090.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1090.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/10900.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1210.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1210.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1210.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1230.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1232.4.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1240.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/12600.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/12600.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/12700.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1300.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1300.15.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1300.4.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1300.7.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/13000.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/13000.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/13000.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/13000.5.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/13000.6.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1304.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1306.11.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1306.13.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1306.12.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1306.14.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1306.15.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1306.16.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1306.17.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1306.18.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1306.19.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1306.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1306.20.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1306.21.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1306.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1306.22.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1306.4.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1306.5.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1306.6.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1306.7.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1306.8.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1306.9.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1307.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1310.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1310.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1320.4.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1400.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/14000.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/14000.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/14100.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/14400.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1450.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/14950.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1510.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1510.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/1520.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2100.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2100.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2100.6.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2200.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2200.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2300.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2410.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2410.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2450.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2450.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2450.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2500.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2500.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2530.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2532.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2610.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2610.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2620.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2620.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2620.5.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2640.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2650.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2660.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2680.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2780.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/2791.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/3100.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/3200.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/3200.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/3230.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/3300.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/3300.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/3300.4.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/3410.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4400.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5100.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5100.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5100.4.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5100.5.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5110.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5110.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5220.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5220.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5300.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5420.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5420.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5500.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5500.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5500.4.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5600.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5610.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/dr-4030-335-002.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/dr-4040-430.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5620.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5710.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5720.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5730.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5720.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5740.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6000.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6020.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6030.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6090.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6100.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6100.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6100.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6100.4.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6100.6.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6100.8.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4339.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4351.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4410.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4410.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4410.5.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4420.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4430.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4430.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4430.5.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4440.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4451.12.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4451.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4451.4.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4451.5.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4451.7.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4451.8.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4451.9.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4461.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4500.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4530.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4531.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4536.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4550.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4550.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4550.4.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4550.7.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4551.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4591.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4610.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4610.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4610.5.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4610.6.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4610.9.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4630.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4630.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4630.5.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4630.7.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4711.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4713.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4713.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4732.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4735.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4735.4.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4735.7.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4735.8.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4735.9.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4771.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4791.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4791.11.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4791.12.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4791.13.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4791.16.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4791.5.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4791.6.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4791.8.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4810.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/4831.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5000.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5000.10.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5000.15.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5000.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5000.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5000.4.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5000.5.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5000.6.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5000.7.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5000.8.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5000.9.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5010.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5020.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5020.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5030.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5030.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5060.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/5090.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6110.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6120.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6170.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6210.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6240.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6300.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6330.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6400.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6410.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6410.4.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6420.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6420.5.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6500.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6600.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6600.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6700.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6900.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/6910.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/7000.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/7000.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/7000.4.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/7010.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/7020.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/7111.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/7120.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/7130.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/7150.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/7160.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/7160.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/7160.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/7221.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/7230.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/7310.5.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/7320.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/7355.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/7520.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/7530.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/7530.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/7620.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/8000.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/8010.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/8010.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/8010.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/8010.4.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/8010.5.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/8021.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/8030.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/8080.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/8080.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/8091.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/8091.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/8140.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/8150.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/8160.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/8410.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9000.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9000.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9000.6.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9000.8.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9000.9.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9010.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9040.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9040.5.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9500.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9500.8.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9500.9.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9510.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9530.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9700.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9770.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9780.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9790.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9900.1.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9900.2.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9900.3.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9900.4.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9900.5.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9900.6.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9900.7.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9900.8.pdf",
        "https://www.fsis.usda.gov/sites/default/files/media_file/2020-07/9910.1.pdf",
    ]


    all_citations = []
    for url in url_list:
        all_citations.extend(process_url(url))
        time.sleep(3)  # pause between downloads to mimic human browsing


    save_to_excel(all_citations)


if __name__ == "__main__":
    main()






